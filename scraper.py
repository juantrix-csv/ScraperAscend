import os
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

BASE_URL = "https://wakomercadonatural.com/tienda/"


def fetch_products(page=1):
    url = BASE_URL
    if page > 1:
        url = f"{BASE_URL}page/{page}/"
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (X11; Linux x86_64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/118.0 Safari/537.36"
        )
    }
    resp = requests.get(url, headers=headers)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    products = []
    seen = set()
    for prod in soup.select("div.product-small"):
        name_el = prod.select_one("div.title-wrapper p.name a")
        price_el = prod.select_one("span.price span.woocommerce-Price-amount")
        if not (name_el and price_el):
            continue
        url = name_el.get("href")
        if url in seen:
            continue
        seen.add(url)
        name = name_el.get_text(strip=True)
        price = price_el.get_text(strip=True)
        products.append({"name": name, "price": price, "url": url})
    return products


def scrape_all_pages(max_pages=1):
    page = 1
    all_products = []
    while True:
        products = fetch_products(page)
        if not products:
            break
        all_products.extend(products)
        page += 1
        if max_pages and page > max_pages:
            break
    return all_products


def save_to_excel(products, filename="productos.xlsx"):
    """Create or update an Excel file with the product information."""
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Nombre", "Precio", "URL"])

    # Map existing URLs to row numbers
    existing = {
        ws.cell(row=row, column=3).value: row
        for row in range(2, ws.max_row + 1)
    }

    for product in products:
        row = existing.get(product["url"])
        if row:
            ws.cell(row=row, column=1, value=product["name"])
            ws.cell(row=row, column=2, value=product["price"])
        else:
            ws.append([product["name"], product["price"], product["url"]])

    wb.save(filename)


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Scrape product prices")
    parser.add_argument(
        "--pages",
        type=int,
        default=1,
        help="Number of pages to scrape (default: 1)",
    )
    args = parser.parse_args()

    products = scrape_all_pages(args.pages)
    save_to_excel(products)
    for product in products:
        print(f"{product['name']} - {product['price']}")


if __name__ == "__main__":
    main()
